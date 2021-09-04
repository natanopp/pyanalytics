## sales report automation
print('Running sales report and AR%')
print()

## import modules
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

## import data
# reference month order
mo_order = pd.DataFrame({'Month': ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                                   'JUL', 'AUG', 'SEP','OCT', 'NOV', 'DEC'],
                         'MonthNo': [i for i in range(1, 13)]})

# reference sales department
depts = ['BKK1', 'BKK2', 'UTH', 'LTH']

# sales report
report = pd.read_csv('data/output_sales_report.csv')

# sales target
targets = pd.DataFrame()
wb = load_workbook(filename='data/Sales_targets.xlsx')
for ws in wb.sheetnames:
    print('Get worksheet:', ws)
    target = pd.read_excel('data/Sales_targets.xlsx', header=4,
                            sheet_name=ws,
                            usecols=[i for i in range(0,13)])
    targets = pd.concat([targets, target])

# sales target processing
mlt_target = pd.melt(targets, id_vars='SalesRep',
                    value_vars=targets.columns[1:],
                    var_name = 'Month', value_name = 'Target')

mlt_target['Target'] = mlt_target['Target'].round(0).astype(int)

# map sales with targets
compare = pd.merge(report, mlt_target,
                  on=['SalesRep', 'Month'], how='inner')

# map month order
compare = pd.merge(compare, mo_order,
                  on=['Month'], how='inner')

# calculate AR%
compare['AR'] = compare['Sales']/compare['Target']

# pivot
pvt_compare = compare.pivot(index=['SalesRep'], columns=['MonthNo'],
                            values=['AR']).reset_index(drop=False)

# write to excel
# create workbook
wb = Workbook()

# iterate through department
print()
for dept in depts:
    print('Creating report:', dept)
    ws = wb.create_sheet(dept)

    # create title and attibutes
    ws['A1'] = 'Target Report'
    ws['A1'].font = Font(bold=True)

    ws['A3'] = 'Team'
    ws['B3'] = dept
    ws['B3'].font = Font(color="00FF0000", italic=True)

    ws['A5'] = 'SalesRep'
    ws['A5'].fill = PatternFill("solid", fgColor="00FFFF00")
    ws['A5'].font = Font(bold=True)

    # create columns
    for row in ws.iter_rows(min_row=5, max_row=5, min_col=2, max_col=13):
        i = 0
        for cell in row:
            cell.value = mo_order['Month'][i]
            cell.fill = PatternFill("solid", fgColor="00FFFF00")
            cell.font = Font(bold=True)
            i+=1

    # load/add data
    # filter
    print('Filtering:', dept)
    print()
    data = pvt_compare[pvt_compare['SalesRep'].str.contains(dept)]
    for r in dataframe_to_rows(data, index=False, header=False):
        ws.append(r)

    # number format - percent
    for row in ws.iter_rows(min_row=6, max_row=6+(len(data)-1),
                            min_col=2, max_col=13):
        for cell in row:
            cell.number_format = '0.00%'

# save to excel
wb.save('data/ar_report.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
wb.remove(sheet)
wb.save('data/ar_report.xlsx')
